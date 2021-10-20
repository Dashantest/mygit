#! /usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2021/3/24 23:32
# @Author : 大山
from openpyxl import load_workbook

def from_excel_case(file, sheet_name=None):
    # 获取工作簿
    wb = load_workbook(file)
    # print(wb)
    # 获取表，类似字典取值的方式
    if sheet_name is None:
        ws = wb.active  # 获取当前表
    else:
        ws = wb[sheet_name]
    # 获取单元格，通过cell()转入行和列
    # ws.cell(row=1,column=1)
    # 获取最大行和最大列
    row = ws.max_row
    column = ws.max_column
    title = {} # 定义空字典获取第一行的title
    data = [] # 定义空列表，接收每一行的数据
    for i in range(1, column+1):
        title[i] = ws.cell(row=1, column=i).value  # 通过字典取值的方式获取第一行每一列的值
    for j in range(2, row+1):
        num = {} # 定义一个空字典接收每一行的数据
        for i in range(1, column+1):
            num[title[i]] = ws.cell(row=j, column=i).value
        data.append(num)
    # print(data)
    return data


if __name__ == '__main__':
    from_excel_case(r'C:\Users\zhengshan\PyProjects\day47\test_data.xlsx')